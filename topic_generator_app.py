#!/usr/bin/env python3
"""
Topic Generator Web Application
Automatically generates MECE topics in Hebrew from open-ended responses using GPT 5.2 via Azure.

Features:
- Upload Excel/CSV files
- Select answer column
- Adjustable max topics (2-15) via slider
- AI-powered MECE topic generation in Hebrew
- Automatic classification of all responses
"""
from __future__ import annotations

import csv
import os
import sys
import webbrowser
import subprocess
import platform
import tempfile
from pathlib import Path
from typing import Tuple, List, Optional, Any, Dict, Union
from threading import Timer
from datetime import datetime
from queue import Queue
import threading
import time

from flask import Flask, render_template, request, jsonify, session, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import requests
import yaml
from dotenv import load_dotenv


app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
# Use a fixed secret key based on machine info for session persistence across restarts
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'topic-generator-app-secret-key-2024')
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours

ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

progress_queues = {}

# Store file info outside of session for reliability
file_store = {}


def get_file_info(session_id=None):
    """Get file info from session or backup store."""
    # Try session first
    if 'uploaded_file' in session:
        return {
            'uploaded_file': session.get('uploaded_file'),
            'original_filename': session.get('original_filename'),
            'current_sheet': session.get('current_sheet'),
            'generated_topics': session.get('generated_topics', []),
            'download_file_path': session.get('download_file_path'),
            'download_filename': session.get('download_filename')
        }
    
    # Try backup store
    sid = session_id or session.get('session_id')
    if sid and sid in file_store:
        return file_store[sid]
    
    return None


def update_file_info(session_id, **kwargs):
    """Update file info in both session and backup store."""
    for key, value in kwargs.items():
        session[key] = value
    
    sid = session_id or session.get('session_id')
    if sid:
        if sid not in file_store:
            file_store[sid] = {}
        file_store[sid].update(kwargs)


# ================================================================
# CONFIGURATION
# ================================================================

class AzureOpenAIConfig:
    """Configuration for Azure OpenAI GPT 5.2 and DeepSeek"""
    
    def __init__(self):
        load_dotenv()
        
        # GPT 5.2 Configuration
        self.api_key = os.getenv("OPENAI_API_KEY") or os.getenv("API_KEY")
        if not self.api_key:
            raise ValueError("Missing OPENAI_API_KEY or API_KEY in environment.")
        
        self.api_base_url = (
            os.getenv("OPENAI_API_BASE_URL")
            or os.getenv("API_BASE_URL")
            or "https://api.openai.com"
        )
        self.model = os.getenv("MODEL") or os.getenv("OPENAI_MODEL", "gpt-5.2")
        self.api_version = os.getenv("AZURE_API_VERSION", "2025-04-01-preview")
        self.max_retries = int(os.getenv("OPENAI_MAX_RETRIES", "5"))
        self.retry_backoff_seconds = float(os.getenv("OPENAI_RETRY_BACKOFF_SECONDS", "2"))
        
        # DeepSeek Configuration
        self.deepseek_api_key = os.getenv("REPAIR_API_KEY") or self.api_key
        self.deepseek_model = os.getenv("REPAIR_MODEL", "DeepSeek-V3.1-gavriel")
        self.deepseek_endpoint = os.getenv("REPAIR_ENDPOINT", "https://sni-ai-foundry.services.ai.azure.com/openai/v1/")
    
    def is_azure(self) -> bool:
        return "azure" in self.api_base_url.lower() or "cognitiveservices" in self.api_base_url.lower()
    
    def get_chat_completions_url(self) -> str:
        base = self.api_base_url.rstrip("/")
        if self.is_azure():
            return f"{base}/openai/deployments/{self.model}/chat/completions?api-version={self.api_version}"
        else:
            return f"{base}/v1/chat/completions"
    
    def get_deepseek_url(self) -> str:
        return f"{self.deepseek_endpoint.rstrip('/')}/chat/completions"


# ================================================================
# UTILITY FUNCTIONS
# ================================================================

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def column_letter_to_index(letter: str) -> int:
    letter = letter.upper().strip()
    result = 0
    for char in letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def column_index_to_letter(index: int) -> str:
    letter = ''
    index += 1
    while index > 0:
        index -= 1
        letter = chr(index % 26 + ord('A')) + letter
        index //= 26
    return letter


def _get_csv_metadata(filepath: Path) -> Tuple[List[str], int]:
    with open(filepath, 'r', encoding='utf-8', errors='replace', newline='') as f:
        reader = csv.reader(f)
        columns = next(reader, [])
        f.seek(0)
        row_count = sum(1 for _ in f) - 1
    return columns, max(0, row_count)


def _get_excel_metadata(filepath: Path) -> Tuple[List[str], dict]:
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    first_sheet = wb[sheet_names[0]]
    first_row = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns = [str(c) if c is not None else f"Column_{i}" for i, c in enumerate(first_row)]
    wb.close()
    return sheet_names, {sheet_names[0]: columns}


def _get_excel_sheet_columns(filepath: Path, sheet_name: str) -> List[str]:
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns = [str(c) if c is not None else f"Column_{i}" for i, c in enumerate(first_row)]
    wb.close()
    return columns


def load_data_for_processing(filepath: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
    extension = filepath.suffix.lower()
    
    if extension == '.csv':
        df = pd.read_csv(filepath, dtype=str, low_memory=False)
    else:
        df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)
    
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]
    return df


def _strip_code_fences(content: str) -> str:
    content = content.strip()
    if content.startswith("```"):
        content = content[3:]
        newline_idx = content.find("\n")
        if newline_idx != -1:
            content = content[newline_idx + 1:]
        if content.endswith("```"):
            content = content[:-3]
    return content.strip()


# ================================================================
# TOPIC GENERATION - MULTI-MODEL PARALLEL APPROACH
# ================================================================

from concurrent.futures import ThreadPoolExecutor, as_completed
import random


def _extract_topics_gpt(
    texts: List[str],
    config: AzureOpenAIConfig,
    max_topics: int,
    min_topics: int,
    existing_topics: List[str] = None
) -> List[str]:
    """Extract topics using GPT 5.2"""
    
    existing_context = ""
    if existing_topics:
        existing_context = f"\n\nTopics already identified (ensure new topics don't overlap with these):\n" + "\n".join(f"- {t}" for t in existing_topics)
    
    enumerated_texts = "\n".join(f"{i+1}. {text[:500]}" for i, text in enumerate(texts))
    
    system_prompt = f"""You are a world-class expert in MECE (Mutually Exclusive, Collectively Exhaustive) topic classification for Hebrew survey responses.

Your task: Identify {min_topics}-{max_topics} STRICTLY MUTUALLY EXCLUSIVE topics from Hebrew open-ended responses.

## CRITICAL: MUTUAL EXCLUSIVITY
Each topic must be COMPLETELY DISTINCT from all others. Before adding a topic, ask:
- "Does this overlap with any other topic I've identified?"
- "Could a response fit into multiple of my topics?"
If YES to either → The topics are NOT mutually exclusive → Redesign them

## Common Mistakes to AVOID:
- Creating "Service quality" AND "Staff behavior" → Staff IS part of service → OVERLAP
- Creating "Pricing" AND "Value for money" → These overlap significantly
- Creating "Wait times" AND "Efficiency" → Wait times IS about efficiency → OVERLAP
- Creating both a general topic and its subtopics

## Rules:
1. MUTUALLY EXCLUSIVE: No overlap between topics. Each response fits ONE topic only.
2. EXHAUSTIVE: Together, topics should cover all possible responses
3. All topics in Hebrew
4. Each topic: detailed descriptive phrase (not single words)
5. {min_topics}-{max_topics} topics total
6. "אחר" (Other) only if absolutely necessary
7. NEVER use curly braces {{ }} or brackets around topics

## Output:
Return ONLY a clean YAML list (no curly braces):
- נושא בעברית
- נושא נוסף בעברית
..."""

    user_prompt = f"""Analyze these Hebrew responses and create {min_topics}-{max_topics} STRICTLY MUTUALLY EXCLUSIVE topics.

IMPORTANT: 
- Ensure NO overlap between topics
- NO curly braces or special brackets in topic names
- Each response should fit into ONE topic only{existing_context}

Responses (Hebrew):
{enumerated_texts}

Return ONLY a clean YAML list of mutually exclusive topics in Hebrew (no curly braces)."""

    url = config.get_chat_completions_url()
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {config.api_key}"
    }
    
    payload = {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
    }
    
    if config.is_azure():
        payload["max_completion_tokens"] = 2000
    else:
        payload["model"] = config.model
        payload["max_tokens"] = 2000
        payload["temperature"] = 0.3
    
    response = requests.post(url, json=payload, headers=headers, timeout=120)
    if response.status_code != 200:
        raise Exception(f"GPT API Error {response.status_code}")
    
    data = response.json()
    content = data["choices"][0]["message"]["content"]
    cleaned = _strip_code_fences(content)
    # Remove any curly braces
    cleaned = cleaned.replace('{', '').replace('}', '')
    topics = yaml.safe_load(cleaned)
    
    if isinstance(topics, list):
        return [str(t).strip().replace('{', '').replace('}', '') for t in topics if t and str(t).strip()]
    return []


def _extract_topics_deepseek(
    texts: List[str],
    config: AzureOpenAIConfig,
    max_topics: int,
    min_topics: int,
    existing_topics: List[str] = None
) -> List[str]:
    """Extract topics using DeepSeek"""
    
    existing_context = ""
    if existing_topics:
        existing_context = f"\n\nTopics already identified (ensure new topics don't overlap with these):\n" + "\n".join(f"- {t}" for t in existing_topics)
    
    enumerated_texts = "\n".join(f"{i+1}. {text[:500]}" for i, text in enumerate(texts))
    
    system_prompt = f"""You are a world-class expert in MECE (Mutually Exclusive, Collectively Exhaustive) topic classification for Hebrew survey responses.

Your task: Identify {min_topics}-{max_topics} STRICTLY MUTUALLY EXCLUSIVE topics from Hebrew open-ended responses.

## CRITICAL: MUTUAL EXCLUSIVITY
Each topic must be COMPLETELY DISTINCT from all others. Before adding a topic, ask:
- "Does this overlap with any other topic I've identified?"
- "Could a response fit into multiple of my topics?"
If YES to either → The topics are NOT mutually exclusive → Redesign them

## Common Mistakes to AVOID:
- Creating "Service quality" AND "Staff behavior" → Staff IS part of service → OVERLAP
- Creating "Pricing" AND "Value for money" → These overlap significantly
- Creating "Wait times" AND "Efficiency" → Wait times IS about efficiency → OVERLAP
- Creating both a general topic and its subtopics

## Rules:
1. MUTUALLY EXCLUSIVE: No overlap between topics. Each response fits ONE topic only.
2. EXHAUSTIVE: Together, topics should cover all possible responses
3. All topics in Hebrew
4. Each topic: detailed descriptive phrase (not single words)
5. {min_topics}-{max_topics} topics total
6. "אחר" (Other) only if absolutely necessary
7. NEVER use curly braces {{ }} or brackets around topics

## Output:
Return ONLY a clean YAML list (no curly braces):
- נושא בעברית
- נושא נוסף בעברית
..."""

    user_prompt = f"""Analyze these Hebrew responses and create {min_topics}-{max_topics} STRICTLY MUTUALLY EXCLUSIVE topics.

IMPORTANT: 
- Ensure NO overlap between topics
- NO curly braces or special brackets in topic names
- Each response should fit into ONE topic only{existing_context}

Responses (Hebrew):
{enumerated_texts}

Return ONLY a clean YAML list of mutually exclusive topics in Hebrew (no curly braces)."""

    url = config.get_deepseek_url()
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {config.deepseek_api_key}"
    }
    
    payload = {
        "model": config.deepseek_model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "max_tokens": 2000,
        "temperature": 0.3
    }
    
    response = requests.post(url, json=payload, headers=headers, timeout=120)
    if response.status_code != 200:
        raise Exception(f"DeepSeek API Error {response.status_code}")
    
    data = response.json()
    content = data["choices"][0]["message"]["content"]
    cleaned = _strip_code_fences(content)
    # Remove any curly braces
    cleaned = cleaned.replace('{', '').replace('}', '')
    topics = yaml.safe_load(cleaned)
    
    if isinstance(topics, list):
        return [str(t).strip().replace('{', '').replace('}', '') for t in topics if t and str(t).strip()]
    return []


def _summarize_topics_gpt(
    accumulated_topics: List[str],
    config: AzureOpenAIConfig,
    max_topics: int,
    min_topics: int
) -> List[str]:
    """GPT summarizes its accumulated topics into a MECE list"""
    
    topics_text = "\n".join(f"- {t}" for t in accumulated_topics)
    
    system_prompt = f"""You are a world-class expert in creating MECE (Mutually Exclusive, Collectively Exhaustive) topic classifications.

You have accumulated topics from analyzing multiple batches of Hebrew survey responses. Now you must consolidate them into a FINAL MECE list of {min_topics}-{max_topics} topics.

## CRITICAL RULES:

### 1. MUTUAL EXCLUSIVITY
Each topic must be COMPLETELY DISTINCT. For any pair of topics, ask:
- "Could ONE response fit into BOTH topics?"
- If YES → MERGE them into one topic

### 2. NO OVERLAPS ALLOWED
Common overlaps to eliminate:
- "Service quality" + "Staff attitude" → Staff IS service → MERGE
- "Pricing" + "Value for money" → OVERLAP → MERGE
- General topic + its subtopic → MERGE

### 3. FORMAT RULES
- NEVER use curly braces {{ or }} in topic names
- Each topic: clear Hebrew phrase (not single words)
- No special characters or brackets around topics
- {min_topics}-{max_topics} topics total

## Output:
Return ONLY a clean YAML list:
- נושא ראשון בעברית
- נושא שני בעברית
..."""

    user_prompt = f"""Consolidate these accumulated topics into {min_topics}-{max_topics} STRICTLY MUTUALLY EXCLUSIVE topics in Hebrew:

{topics_text}

IMPORTANT:
1. Merge any overlapping topics
2. NO curly braces or special brackets in topic names
3. Each response should fit into ONE topic only

Return ONLY a clean YAML list of mutually exclusive topics in Hebrew."""

    url = config.get_chat_completions_url()
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {config.api_key}"
    }
    
    payload = {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
    }
    
    if config.is_azure():
        payload["max_completion_tokens"] = 2000
    else:
        payload["model"] = config.model
        payload["max_tokens"] = 2000
        payload["temperature"] = 0.2
    
    response = requests.post(url, json=payload, headers=headers, timeout=120)
    if response.status_code != 200:
        raise Exception(f"GPT Summarize API Error {response.status_code}")
    
    data = response.json()
    content = data["choices"][0]["message"]["content"]
    cleaned = _strip_code_fences(content)
    # Remove any curly braces from topics
    cleaned = cleaned.replace('{', '').replace('}', '')
    topics = yaml.safe_load(cleaned)
    
    if isinstance(topics, list):
        # Clean each topic of curly braces
        return [str(t).strip().replace('{', '').replace('}', '') for t in topics if t and str(t).strip()]
    return []


def _summarize_topics_deepseek(
    accumulated_topics: List[str],
    config: AzureOpenAIConfig,
    max_topics: int,
    min_topics: int
) -> List[str]:
    """DeepSeek summarizes its accumulated topics into a MECE list"""
    
    topics_text = "\n".join(f"- {t}" for t in accumulated_topics)
    
    system_prompt = f"""You are a world-class expert in creating MECE (Mutually Exclusive, Collectively Exhaustive) topic classifications.

You have accumulated topics from analyzing multiple batches of Hebrew survey responses. Now you must consolidate them into a FINAL MECE list of {min_topics}-{max_topics} topics.

## CRITICAL RULES:

### 1. MUTUAL EXCLUSIVITY
Each topic must be COMPLETELY DISTINCT. For any pair of topics, ask:
- "Could ONE response fit into BOTH topics?"
- If YES → MERGE them into one topic

### 2. NO OVERLAPS ALLOWED
Common overlaps to eliminate:
- "Service quality" + "Staff attitude" → Staff IS service → MERGE
- "Pricing" + "Value for money" → OVERLAP → MERGE
- General topic + its subtopic → MERGE

### 3. FORMAT RULES
- NEVER use curly braces {{ or }} in topic names
- Each topic: clear Hebrew phrase (not single words)
- No special characters or brackets around topics
- {min_topics}-{max_topics} topics total

## Output:
Return ONLY a clean YAML list:
- נושא ראשון בעברית
- נושא שני בעברית
..."""

    user_prompt = f"""Consolidate these accumulated topics into {min_topics}-{max_topics} STRICTLY MUTUALLY EXCLUSIVE topics in Hebrew:

{topics_text}

IMPORTANT:
1. Merge any overlapping topics
2. NO curly braces or special brackets in topic names
3. Each response should fit into ONE topic only

Return ONLY a clean YAML list of mutually exclusive topics in Hebrew."""

    url = config.get_deepseek_url()
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {config.deepseek_api_key}"
    }
    
    payload = {
        "model": config.deepseek_model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "max_tokens": 2000,
        "temperature": 0.2
    }
    
    response = requests.post(url, json=payload, headers=headers, timeout=120)
    if response.status_code != 200:
        raise Exception(f"DeepSeek Summarize API Error {response.status_code}")
    
    data = response.json()
    content = data["choices"][0]["message"]["content"]
    cleaned = _strip_code_fences(content)
    # Remove any curly braces from topics
    cleaned = cleaned.replace('{', '').replace('}', '')
    topics = yaml.safe_load(cleaned)
    
    if isinstance(topics, list):
        # Clean each topic of curly braces
        return [str(t).strip().replace('{', '').replace('}', '') for t in topics if t and str(t).strip()]
    return []


def _judge_final_topics(
    gpt_summarized: List[str],
    deepseek_summarized: List[str],
    config: AzureOpenAIConfig,
    max_topics: int,
    min_topics: int
) -> List[str]:
    """GPT acts as final judge to create the ultimate MECE topic list from both summarized lists"""
    
    topics_a = "\n".join(f"- {t}" for t in gpt_summarized)
    topics_b = "\n".join(f"- {t}" for t in deepseek_summarized)
    
    system_prompt = f"""You are the FINAL JUDGE for creating a MECE (Mutually Exclusive, Collectively Exhaustive) topic classification system.

Two independent analysts have each summarized their findings into MECE topic lists from the same Hebrew survey responses. Your job: Create the ULTIMATE final list.

## YOUR CRITICAL TASK:

### 1. STRICT MUTUAL EXCLUSIVITY
For EVERY pair of topics in your final list, verify:
- "Could a SINGLE response fit into BOTH of these?"
- If YES → They MUST be merged

### 2. ELIMINATE ALL OVERLAPS
Even if topics come from different analysts, they may overlap:
- Similar concepts with different wording → MERGE
- General + specific versions → Keep ONE
- Cause and effect relationships → MERGE
- Synonymous descriptions → MERGE

### 3. HONOR THE LIMIT
User requested MAXIMUM {max_topics} topics. Do NOT exceed this.

### 4. FORMAT RULES (CRITICAL)
- NEVER include curly braces {{ or }} in any topic
- NEVER include brackets [ ] or parentheses with special meaning
- Each topic: clean Hebrew phrase
- No decorations or special characters around topic text

## Output:
Return ONLY a clean YAML list of EXACTLY {min_topics}-{max_topics} topics:
- נושא בעברית
- נושא נוסף בעברית
..."""

    user_prompt = f"""Two analysts produced these MECE topic summaries from Hebrew responses:

Analyst A's topics:
{topics_a}

Analyst B's topics:
{topics_b}

Create the FINAL list of {min_topics}-{max_topics} MUTUALLY EXCLUSIVE topics in Hebrew.

CRITICAL CHECKLIST:
✓ Maximum {max_topics} topics (user's limit)
✓ NO overlapping topics - merge any that could apply to same response
✓ NO curly braces or special brackets in topic names
✓ All topics in Hebrew

Return ONLY a clean YAML list."""

    url = config.get_chat_completions_url()
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {config.api_key}"
    }
    
    payload = {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
    }
    
    if config.is_azure():
        payload["max_completion_tokens"] = 2000
    else:
        payload["model"] = config.model
        payload["max_tokens"] = 2000
        payload["temperature"] = 0.2
    
    response = requests.post(url, json=payload, headers=headers, timeout=120)
    if response.status_code != 200:
        raise Exception(f"Judge API Error {response.status_code}")
    
    data = response.json()
    content = data["choices"][0]["message"]["content"]
    cleaned = _strip_code_fences(content)
    topics = yaml.safe_load(cleaned)
    
    if isinstance(topics, list):
        return [str(t).strip() for t in topics if t and str(t).strip()][:max_topics]
    return []


def generate_mece_topics(
    texts: List[str],
    config: AzureOpenAIConfig,
    max_topics: int = 10,
    min_topics: int = 2,
    progress_callback=None
) -> List[str]:
    """
    Generate MECE topics in Hebrew using DUAL-MODEL PARALLEL approach.
    
    Process:
    1. Split texts into batches
    2. For each batch, run GPT 5.2 and DeepSeek in parallel
    3. Each batch receives accumulated topics from previous batches as context
    4. After all batches, GPT 5.2 acts as judge to create final topic list
    
    Args:
        texts: List of response texts to analyze
        config: Azure OpenAI configuration
        max_topics: Maximum number of topics (2-15)
        min_topics: Minimum number of topics (2)
        progress_callback: Optional callback for progress updates
    
    Returns:
        List of Hebrew topic names
    """
    max_topics = max(min_topics, min(15, max_topics))
    
    # Filter valid texts
    valid_texts = [t for t in texts if t and len(t.strip()) > 2]
    
    if not valid_texts:
        raise ValueError("No valid texts found for topic generation")
    
    # Shuffle and limit to reasonable sample
    random.shuffle(valid_texts)
    sample_texts = valid_texts[:min(150, len(valid_texts))]
    
    if progress_callback:
        progress_callback(f"🔄 Starting dual-model topic extraction on {len(sample_texts)} responses...")
    
    # Split into small batches (similar to open_questions app)
    # Smaller batches avoid information overload and allow better nuance detection
    batch_size = 8  # Small batches for better topic quality
    batches = [sample_texts[i:i+batch_size] for i in range(0, len(sample_texts), batch_size)]
    
    if progress_callback:
        progress_callback(f"📦 Split into {len(batches)} batches ({batch_size} texts each) for parallel processing")
    
    # Accumulated topics from each model
    gpt_accumulated_topics: List[str] = []
    deepseek_accumulated_topics: List[str] = []
    
    # Process each batch
    for batch_idx, batch_texts in enumerate(batches, 1):
        if progress_callback:
            progress_callback(f"⚡ Processing batch {batch_idx}/{len(batches)} with GPT 5.2 + DeepSeek in parallel...")
        
        gpt_result = None
        deepseek_result = None
        gpt_error = None
        deepseek_error = None
        
        # Run both models in parallel
        with ThreadPoolExecutor(max_workers=2) as executor:
            gpt_future = executor.submit(
                _extract_topics_gpt,
                batch_texts,
                config,
                max_topics,
                min_topics,
                gpt_accumulated_topics if gpt_accumulated_topics else None
            )
            deepseek_future = executor.submit(
                _extract_topics_deepseek,
                batch_texts,
                config,
                max_topics,
                min_topics,
                deepseek_accumulated_topics if deepseek_accumulated_topics else None
            )
            
            try:
                gpt_result = gpt_future.result(timeout=180)
                if progress_callback:
                    progress_callback(f"  ✅ GPT 5.2: found {len(gpt_result)} topics")
            except Exception as e:
                gpt_error = str(e)
                if progress_callback:
                    progress_callback(f"  ⚠️ GPT 5.2 error: {gpt_error}")
            
            try:
                deepseek_result = deepseek_future.result(timeout=180)
                if progress_callback:
                    progress_callback(f"  ✅ DeepSeek: found {len(deepseek_result)} topics")
            except Exception as e:
                deepseek_error = str(e)
                if progress_callback:
                    progress_callback(f"  ⚠️ DeepSeek error: {deepseek_error}")
        
        # Accumulate topics (merge with existing, removing exact duplicates)
        if gpt_result:
            for topic in gpt_result:
                if topic not in gpt_accumulated_topics:
                    gpt_accumulated_topics.append(topic)
        
        if deepseek_result:
            for topic in deepseek_result:
                if topic not in deepseek_accumulated_topics:
                    deepseek_accumulated_topics.append(topic)
        
        # Small delay between batches
        if batch_idx < len(batches):
            time.sleep(0.5)
    
    if progress_callback:
        progress_callback(f"📊 GPT accumulated {len(gpt_accumulated_topics)} raw topics")
        progress_callback(f"📊 DeepSeek accumulated {len(deepseek_accumulated_topics)} raw topics")
    
    # Handle case where one model completely failed
    if not gpt_accumulated_topics and not deepseek_accumulated_topics:
        raise RuntimeError("Both models failed to generate any topics")
    
    # === FINAL THREE-STEP PROCESS ===
    
    # Step 1: DeepSeek summarizes its accumulated topics into MECE list
    deepseek_summarized = []
    if deepseek_accumulated_topics:
        if progress_callback:
            progress_callback("🔄 Step 1/3: DeepSeek summarizing its topics into MECE list...")
        try:
            deepseek_summarized = _summarize_topics_deepseek(
                deepseek_accumulated_topics, config, max_topics, min_topics
            )
            if progress_callback:
                progress_callback(f"  ✅ DeepSeek summarized to {len(deepseek_summarized)} MECE topics")
        except Exception as e:
            if progress_callback:
                progress_callback(f"  ⚠️ DeepSeek summarize error: {e}")
            deepseek_summarized = deepseek_accumulated_topics[:max_topics]
    
    # Step 2: GPT summarizes its accumulated topics into MECE list
    gpt_summarized = []
    if gpt_accumulated_topics:
        if progress_callback:
            progress_callback("🔄 Step 2/3: GPT summarizing its topics into MECE list...")
        try:
            gpt_summarized = _summarize_topics_gpt(
                gpt_accumulated_topics, config, max_topics, min_topics
            )
            if progress_callback:
                progress_callback(f"  ✅ GPT summarized to {len(gpt_summarized)} MECE topics")
        except Exception as e:
            if progress_callback:
                progress_callback(f"  ⚠️ GPT summarize error: {e}")
            gpt_summarized = gpt_accumulated_topics[:max_topics]
    
    # Handle single model success
    if not gpt_summarized and deepseek_summarized:
        if progress_callback:
            progress_callback("⚠️ Using DeepSeek summary only (GPT failed)")
        return [t.replace('{', '').replace('}', '') for t in deepseek_summarized[:max_topics]]
    
    if not deepseek_summarized and gpt_summarized:
        if progress_callback:
            progress_callback("⚠️ Using GPT summary only (DeepSeek failed)")
        return [t.replace('{', '').replace('}', '') for t in gpt_summarized[:max_topics]]
    
    # Step 3: GPT (as judge) creates final MECE list from both summaries
    if progress_callback:
        progress_callback("⚖️ Step 3/3: GPT judging final topics from both summaries...")
    
    try:
        final_topics = _judge_final_topics(
            gpt_summarized,
            deepseek_summarized,
            config,
            max_topics,
            min_topics
        )
        
        # Clean any remaining curly braces
        final_topics = [t.replace('{', '').replace('}', '') for t in final_topics]
        
        if final_topics and len(final_topics) >= min_topics:
            if progress_callback:
                progress_callback(f"✅ Final MECE topics: {len(final_topics)} topics selected (max was {max_topics})")
            return final_topics[:max_topics]
        else:
            # Fallback: merge both summarized lists
            if progress_callback:
                progress_callback("⚠️ Judge produced insufficient topics, merging summaries...")
            merged = []
            for t in gpt_summarized + deepseek_summarized:
                clean_t = t.replace('{', '').replace('}', '')
                if clean_t not in merged:
                    merged.append(clean_t)
            return merged[:max_topics]
            
    except Exception as e:
        if progress_callback:
            progress_callback(f"⚠️ Judging error: {e}, merging summaries...")
        merged = []
        for t in gpt_summarized + deepseek_summarized:
            clean_t = t.replace('{', '').replace('}', '')
            if clean_t not in merged:
                merged.append(clean_t)
        return merged[:max_topics]


# ================================================================
# CLASSIFICATION
# ================================================================

def classify_texts_batch(
    texts: List[str],
    topics: List[str],
    config: AzureOpenAIConfig,
    progress_callback=None,
    batch_size: int = 5
) -> List[Dict[str, int]]:
    """
    Classify a batch of texts against the generated topics.
    
    Returns: List of dicts mapping topic -> 0/1
    """
    results = []
    total_batches = (len(texts) + batch_size - 1) // batch_size
    
    for batch_idx in range(0, len(texts), batch_size):
        batch_texts = texts[batch_idx:batch_idx + batch_size]
        current_batch = batch_idx // batch_size + 1
        
        if progress_callback:
            progress_callback(f"Classifying batch {current_batch}/{total_batches} ({len(batch_texts)} texts)...")
        
        batch_results = _classify_batch_internal(batch_texts, topics, config)
        results.extend(batch_results)
        
        # Small delay between batches
        if batch_idx + batch_size < len(texts):
            time.sleep(0.3)
    
    return results


def _classify_batch_internal(
    texts: List[str],
    topics: List[str],
    config: AzureOpenAIConfig
) -> List[Dict[str, int]]:
    """Internal batch classification using GPT 5.2"""
    
    # Handle empty texts
    valid_texts = []
    text_indices = []
    for i, text in enumerate(texts):
        if text and len(text.strip()) > 2:
            valid_texts.append(text)
            text_indices.append(i)
    
    # Initialize results with all 0s
    results = [{topic: 0 for topic in topics} for _ in texts]
    
    if not valid_texts:
        return results
    
    topics_text = ", ".join(topics)
    enumerated_texts = "\n\n".join(f"{i+1}. {text}" for i, text in enumerate(valid_texts))
    
    system_prompt = """אתה מומחה בסיווג תוכן בעברית.

המשימה: לסווג כל טקסט לפי הנושאים הנתונים.

כללים:
1. לכל טקסט ולכל נושא, החזר 1 אם הנושא מופיע בבירור בטקסט, אחרת 0
2. טקסט יכול להתאים למספר נושאים
3. אם הטקסט ריק או לא ברור, החזר 0 לכל הנושאים

פורמט הפלט (YAML בלבד):
- id: 1
  נושא א: 0
  נושא ב: 1
- id: 2
  נושא א: 1
  נושא ב: 0"""

    user_prompt = f"""נושאים לסיווג: {topics_text}

טקסטים לסיווג:
{enumerated_texts}

החזר YAML בלבד."""

    url = config.get_chat_completions_url()
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {config.api_key}"
    }
    
    payload = {
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
    }
    
    if config.is_azure():
        payload["max_completion_tokens"] = 3000
    else:
        payload["model"] = config.model
        payload["max_tokens"] = 3000
        payload["temperature"] = 0
    
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=120)
        response.raise_for_status()
        data = response.json()
        
        content = data["choices"][0]["message"]["content"]
        cleaned = _strip_code_fences(content)
        parsed = yaml.safe_load(cleaned)
        
        # Parse results
        if isinstance(parsed, list):
            for item in parsed:
                if not isinstance(item, dict):
                    continue
                text_id = item.get('id')
                if text_id is None:
                    continue
                
                idx = int(text_id) - 1
                if 0 <= idx < len(valid_texts):
                    original_idx = text_indices[idx]
                    for topic in topics:
                        val = item.get(topic, 0)
                        results[original_idx][topic] = 1 if val in (1, "1", True) else 0
        
    except Exception as e:
        print(f"Classification error: {e}", file=sys.stderr)
    
    return results


# ================================================================
# PROGRESS CAPTURE
# ================================================================

class ProgressCapture:
    def __init__(self, session_id: str, queue: Queue):
        self.session_id = session_id
        self.queue = queue
        self.terminal = sys.stdout
        
    def write(self, message):
        if message.strip():
            try:
                self.queue.put(message.strip())
            except Exception:
                pass
        self.terminal.write(message)
    
    def flush(self):
        self.terminal.flush()


# ================================================================
# FLASK ROUTES
# ================================================================

@app.route('/')
def index():
    return render_template('topic_generator.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Please upload .xlsx, .xls, or .csv file'}), 400
    
    try:
        filename = secure_filename(file.filename)
        session_id = os.urandom(16).hex()
        session['session_id'] = session_id
        
        # Use temporary file that will be cleaned up when session ends
        extension = Path(filename).suffix.lower()
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=extension)
        file.save(temp_file.name)
        temp_file.close()
        
        input_path = Path(temp_file.name)
        
        if extension == '.csv':
            columns, row_count = _get_csv_metadata(input_path)
            sheet_names = ['Sheet1']
            default_sheet = 'Sheet1'
        else:
            sheet_names, columns_by_sheet = _get_excel_metadata(input_path)
            default_sheet = sheet_names[0]
            columns = columns_by_sheet.get(default_sheet, [])
            row_count = None
        
        session['current_sheet'] = default_sheet
        
        column_info = [
            {
                'index': i,
                'letter': column_index_to_letter(i),
                'name': str(col)
            }
            for i, col in enumerate(columns)
        ]
        
        session['uploaded_file'] = str(input_path)
        session['original_filename'] = filename
        session['sheet_names'] = sheet_names
        session.permanent = True  # Make session permanent
        
        # Also store in backup file_store for reliability
        file_store[session_id] = {
            'uploaded_file': str(input_path),
            'original_filename': filename,
            'sheet_names': sheet_names,
            'current_sheet': default_sheet,
            'generated_topics': [],
            'download_file_path': None,
            'download_filename': None
        }
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'filename': filename,
            'sheets': sheet_names,
            'default_sheet': default_sheet,
            'columns': column_info,
            'num_rows': row_count,
            'num_columns': len(columns)
        })
    
    except Exception as e:
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500


@app.route('/reset_upload', methods=['POST'])
def reset_upload():
    try:
        filepath_str = session.get('uploaded_file')
        
        if filepath_str:
            filepath = Path(filepath_str)
            if filepath.exists():
                filepath.unlink()
        
        for key in ('uploaded_file', 'original_filename', 'sheet_names', 'current_sheet', 'session_id'):
            session.pop(key, None)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/load_sheet', methods=['POST'])
def load_sheet():
    try:
        data = request.json
        sheet_name = data.get('sheet_name')
        
        if 'session_id' not in session or 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        filepath = Path(session['uploaded_file'])
        
        if not filepath.exists():
            return jsonify({'error': 'Uploaded file not found'}), 400
        
        columns = _get_excel_sheet_columns(filepath, sheet_name)
        session['current_sheet'] = sheet_name
        
        column_info = [
            {
                'index': i,
                'letter': column_index_to_letter(i),
                'name': str(col)
            }
            for i, col in enumerate(columns)
        ]
        
        return jsonify({
            'success': True,
            'sheet_name': sheet_name,
            'columns': column_info,
            'num_columns': len(columns)
        })
    
    except Exception as e:
        return jsonify({'error': f'Error loading sheet: {str(e)}'}), 500


@app.route('/preview', methods=['POST'])
def preview_data():
    try:
        data = request.json
        answer_col = data.get('answer_column')
        
        if 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        filepath = Path(session['uploaded_file'])
        current_sheet = session.get('current_sheet')
        
        if not filepath.exists():
            return jsonify({'error': 'File not found'}), 400
        
        extension = filepath.suffix.lower()
        
        if extension == '.csv':
            df_preview = pd.read_csv(filepath, nrows=5, dtype=str, low_memory=False)
        else:
            df_preview = pd.read_excel(filepath, sheet_name=current_sheet, nrows=5, dtype=str)
        
        df_preview.columns = [c.strip() if isinstance(c, str) else c for c in df_preview.columns]
        
        answer_idx = column_letter_to_index(answer_col)
        
        if answer_idx >= len(df_preview.columns):
            return jsonify({'error': 'Invalid column selection'}), 400
        
        answer_column_name = df_preview.columns[answer_idx]
        
        preview_data = []
        for i in range(len(df_preview)):
            answer_text = str(df_preview.iloc[i, answer_idx])
            if pd.isna(df_preview.iloc[i, answer_idx]):
                answer_text = "(empty)"
            row_data = {
                'row_num': i + 1,
                'answer': answer_text[:150] + '...' if len(answer_text) > 150 else answer_text
            }
            preview_data.append(row_data)
        
        return jsonify({
            'success': True,
            'answer_column': answer_column_name,
            'preview': preview_data
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/generate_topics', methods=['POST'])
def generate_topics():
    """Generate MECE topics from the answer column."""
    try:
        data = request.json
        answer_col = data.get('answer_column')
        max_topics = int(data.get('max_topics', 10))
        
        if 'uploaded_file' not in session:
            return jsonify({'error': 'No file uploaded'}), 400
        
        session_id = session.get('session_id')
        filepath = Path(session['uploaded_file'])
        current_sheet = session.get('current_sheet')
        original_filename = session.get('original_filename', 'output')
        
        # Initialize progress queue
        progress_queues[session_id] = Queue()
        queue = progress_queues[session_id]
        
        answer_idx = column_letter_to_index(answer_col)
        
        def run_topic_generation():
            try:
                queue.put("Loading data...")
                
                config = AzureOpenAIConfig()
                queue.put(f"Using model: {config.model}")
                
                df = load_data_for_processing(filepath, current_sheet)
                answer_column_name = df.columns[answer_idx]
                
                # Get all non-empty answers
                texts = df[answer_column_name].dropna().astype(str).tolist()
                texts = [t for t in texts if t.strip() and len(t.strip()) > 2]
                
                queue.put(f"Found {len(texts)} valid responses")
                
                # Generate topics
                topics = generate_mece_topics(
                    texts,
                    config,
                    max_topics=max_topics,
                    min_topics=2,
                    progress_callback=lambda msg: queue.put(msg)
                )
                
                # Pre-generate the output file so download is instant
                queue.put("📄 Preparing file for download...")
                
                # Remove trailing empty rows - more robust check
                def is_row_empty(row):
                    for val in row:
                        if pd.notna(val):
                            str_val = str(val).strip()
                            if str_val and str_val.lower() not in ('nan', 'none', ''):
                                return False
                    return True
                
                while len(df) > 0:
                    if is_row_empty(df.iloc[-1]):
                        df = df.iloc[:-1]
                    else:
                        break
                
                # Insert topic columns right after the answer column
                answer_col_position = df.columns.get_loc(answer_column_name) + 1
                for i, topic in enumerate(topics):
                    if topic not in df.columns:
                        df.insert(answer_col_position + i, topic, '')
                
                # Create the CSV content and save to temp file
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                original_name = Path(original_filename).stem
                output_filename = f"{original_name}_with_topics_{timestamp}.csv"
                
                # Save to a temp file for quick download
                output_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='wb')
                output_temp.write(b'\xef\xbb\xbf')  # UTF-8 BOM for Excel
                csv_string = df.to_csv(index=False, encoding='utf-8')
                output_temp.write(csv_string.encode('utf-8'))
                output_temp.close()
                
                # Store file info for download
                queue.put(f"FILE_READY:{output_temp.name}|{output_filename}")
                
                # Use JSON encoding to safely pass topics (avoids comma splitting issues)
                import json
                queue.put(f"TOPICS_GENERATED:{json.dumps(topics, ensure_ascii=False)}")
                queue.put("DONE")
                
            except Exception as e:
                import traceback
                error_msg = f"❌ ERROR: {str(e)}\n{traceback.format_exc()}"
                queue.put(error_msg)
                queue.put("ERROR")
        
        thread = threading.Thread(target=run_topic_generation)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'success': True,
            'message': 'Topic generation started',
            'session_id': session_id
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/prepare_download', methods=['POST'])
def prepare_download():
    """Prepare/regenerate the download file with current topics."""
    try:
        data = request.json
        topics = data.get('topics', [])
        answer_column = data.get('answer_column', '')
        req_session_id = data.get('session_id')

        if not topics:
            return jsonify({'error': 'No topics provided'}), 400

        # Get file info from session or backup store
        file_info = get_file_info(req_session_id)
        if not file_info or not file_info.get('uploaded_file'):
            return jsonify({'error': 'No file uploaded. Please upload a file and generate topics again.'}), 400

        filepath = Path(file_info['uploaded_file'])
        current_sheet = file_info.get('current_sheet')
        original_filename = file_info.get('original_filename', 'output')

        if not filepath.exists():
            return jsonify({'error': 'Uploaded file not found. Please upload the file again.'}), 400

        # Load the data
        df = load_data_for_processing(filepath, current_sheet)
        answer_idx = column_letter_to_index(answer_column)

        if answer_idx >= len(df.columns):
            return jsonify({'error': 'Invalid answer column'}), 400

        answer_column_name = df.columns[answer_idx]

        # Remove any previously added topic columns
        old_topics = file_info.get('generated_topics', [])
        cols_to_keep = []
        for col in df.columns:
            if col not in old_topics:
                cols_to_keep.append(col)
        df = df[cols_to_keep]
        
        # Remove trailing empty rows
        def is_row_empty(row):
            for val in row:
                if pd.notna(val):
                    str_val = str(val).strip()
                    if str_val and str_val.lower() not in ('nan', 'none', ''):
                        return False
            return True
        
        while len(df) > 0:
            if is_row_empty(df.iloc[-1]):
                df = df.iloc[:-1]
            else:
                break
        
        # Insert new topic columns right after the answer column
        answer_col_position = df.columns.get_loc(answer_column_name) + 1
        for i, topic in enumerate(topics):
            if topic not in df.columns:
                df.insert(answer_col_position + i, topic, '')
        
        # Create the CSV file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        original_name = Path(original_filename).stem
        output_filename = f"{original_name}_with_topics_{timestamp}.csv"
        
        output_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='wb')
        output_temp.write(b'\xef\xbb\xbf')  # UTF-8 BOM for Excel
        csv_string = df.to_csv(index=False, encoding='utf-8')
        output_temp.write(csv_string.encode('utf-8'))
        output_temp.close()
        
        # Update both session and backup store
        sid = req_session_id or session.get('session_id')
        update_file_info(sid,
            generated_topics=topics,
            download_file_path=output_temp.name,
            download_filename=output_filename
        )
        
        return jsonify({
            'success': True,
            'message': 'File prepared successfully',
            'file_ready': {
                'path': output_temp.name,
                'filename': output_filename
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Prepare download error: {e}\n{traceback.format_exc()}", file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/download_file/<session_id>')
def download_file_route(session_id):
    """Download file directly to Desktop."""
    try:
        # Get file info from session or backup store
        file_info = get_file_info(session_id)
        
        if not file_info:
            return jsonify({'error': 'Session not found. Please upload a file and generate topics again.'}), 400
        
        download_path = file_info.get('download_file_path')
        download_filename = file_info.get('download_filename')
        
        if download_path and Path(download_path).exists():
            # Save directly to Desktop
            desktop_path = Path.home() / "Desktop" / download_filename
            
            # Copy file to desktop
            import shutil
            shutil.copy2(download_path, desktop_path)
            
            return jsonify({
                'success': True,
                'message': f'File saved to Desktop: {download_filename}',
                'path': str(desktop_path)
            })
        
        # Fallback: file not pre-generated, return error
        return jsonify({'error': 'File not ready. Please generate topics first.'}), 400
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/store_topics', methods=['POST'])
def store_topics():
    """Store generated topics and file info in session and backup store."""
    try:
        data = request.json
        topics = data.get('topics', [])
        file_path = data.get('file_path')
        file_name = data.get('file_name')
        req_session_id = data.get('session_id')

        if not topics:
            return jsonify({'error': 'No topics provided'}), 400

        # Update both session and backup store
        sid = req_session_id or session.get('session_id')
        updates = {'generated_topics': topics}
        
        if file_path and file_name:
            updates['download_file_path'] = file_path
            updates['download_filename'] = file_name
        
        update_file_info(sid, **updates)

        return jsonify({'success': True, 'message': f'Stored {len(topics)} topics'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/progress/<session_id>')
def get_progress(session_id):
    if session_id not in progress_queues:
        return jsonify({'messages': [], 'done': False})
    
    messages = []
    done = False
    error = False
    topics = None
    file_ready = None
    
    queue = progress_queues[session_id]
    while not queue.empty():
        msg = queue.get()
        if msg == "DONE":
            done = True
        elif msg == "ERROR":
            done = True
            error = True
        elif msg.startswith("TOPICS_GENERATED:"):
            import json
            topics = json.loads(msg.replace("TOPICS_GENERATED:", ""))
        elif msg.startswith("FILE_READY:"):
            # Format: FILE_READY:path|filename
            file_info = msg.replace("FILE_READY:", "")
            parts = file_info.split("|")
            if len(parts) == 2:
                file_ready = {'path': parts[0], 'filename': parts[1]}
        else:
            messages.append(msg)
    
    response = {
        'messages': messages,
        'done': done,
        'error': error
    }
    
    if topics:
        response['topics'] = topics
    
    if file_ready:
        response['file_ready'] = file_ready
    
    return jsonify(response)


@app.route('/status')
def status():
    try:
        config = AzureOpenAIConfig()
        return jsonify({
            'status': 'ready',
            'model': config.model,
            'api_base_url': config.api_base_url
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e)
        }), 500


@app.route('/chat_feedback', methods=['POST'])
def chat_feedback():
    """Process user feedback to refine topics using GPT 5.2."""
    try:
        data = request.json
        user_message = data.get('message', '')
        current_topics = data.get('current_topics', [])
        answer_column = data.get('answer_column', '')
        
        if not user_message:
            return jsonify({'error': 'No message provided'}), 400
        
        if not current_topics:
            return jsonify({'error': 'No current topics provided'}), 400
        
        config = AzureOpenAIConfig()
        
        # Get sample answers for context
        sample_answers = []
        if 'uploaded_file' in session:
            filepath = Path(session['uploaded_file'])
            current_sheet = session.get('current_sheet')
            if filepath.exists():
                try:
                    df = load_data_for_processing(filepath, current_sheet)
                    answer_idx = column_letter_to_index(answer_column)
                    if answer_idx < len(df.columns):
                        answer_column_name = df.columns[answer_idx]
                        texts = df[answer_column_name].dropna().astype(str).tolist()
                        texts = [t for t in texts if t.strip() and len(t.strip()) > 2]
                        # Get a sample of answers for context
                        import random
                        sample_size = min(20, len(texts))
                        sample_answers = random.sample(texts, sample_size) if texts else []
                except Exception as e:
                    print(f"Error loading sample answers: {e}", file=sys.stderr)
        
        # Build context for the model
        topics_text = "\n".join(f"{i+1}. {topic}" for i, topic in enumerate(current_topics))
        sample_text = "\n".join(f"- {ans[:200]}" for ans in sample_answers[:10]) if sample_answers else "(No samples available)"
        
        system_prompt = """You are an expert assistant specializing in topic analysis and improving MECE (Mutually Exclusive, Collectively Exhaustive) classifications for Hebrew text.

The user has created a list of topics from Hebrew open-ended responses and now wants to refine it.

Your role:
1. Understand the user's request
2. Make ONLY the changes the user requested - no more, no less
3. Return the updated list of topics

Rules:
- If the user asks to merge topics - merge only those specified
- If the user asks to add a topic - add only that one
- If the user asks to change wording - change only the requested phrasing
- Do NOT modify topics the user didn't mention
- Keep all topics in proper Hebrew
- Maintain detailed, descriptive topic names (not single words)

Response Format:
1. First, a brief explanation (one or two sentences) of what you did
2. Then return the updated list in YAML format:

Explanation: [brief explanation in Hebrew]

Topics:
- First topic in Hebrew
- Second topic in Hebrew
..."""

        user_prompt = f"""Current topics (in Hebrew):
{topics_text}

Sample responses from the original data (for context, in Hebrew):
{sample_text}

User request: {user_message}

Make ONLY the requested changes and return the updated list. Keep all topics in Hebrew."""

        url = config.get_chat_completions_url()
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {config.api_key}"
        }
        
        payload = {
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
        }
        
        if config.is_azure():
            payload["max_completion_tokens"] = 1500
        else:
            payload["model"] = config.model
            payload["max_tokens"] = 1500
            payload["temperature"] = 0.3
        
        response = requests.post(url, json=payload, headers=headers, timeout=120)
        
        if response.status_code != 200:
            return jsonify({'error': f'API Error: {response.status_code}'}), 500
        
        api_data = response.json()
        content = api_data["choices"][0]["message"]["content"]
        
        # Parse the response
        explanation = ""
        new_topics = []
        
        # Try to extract explanation and topics (handle both Hebrew and English keywords)
        topics_marker = None
        for marker in ["Topics:", "נושאים:", "topics:"]:
            if marker in content:
                topics_marker = marker
                break
        
        if topics_marker:
            parts = content.split(topics_marker)
            explanation_part = parts[0]
            topics_part = parts[1] if len(parts) > 1 else ""
            
            # Extract explanation (handle both Hebrew and English)
            for expl_marker in ["Explanation:", "הסבר:", "explanation:"]:
                if expl_marker in explanation_part:
                    explanation = explanation_part.split(expl_marker)[-1].strip()
                    break
            else:
                explanation = explanation_part.strip()
            
            # Parse topics YAML
            cleaned_topics = _strip_code_fences(topics_part)
            try:
                parsed_topics = yaml.safe_load(cleaned_topics)
                if isinstance(parsed_topics, list):
                    new_topics = [str(t).strip() for t in parsed_topics if t and str(t).strip()]
            except:
                # Try line-by-line parsing
                for line in cleaned_topics.split('\n'):
                    line = line.strip()
                    if line.startswith('- '):
                        topic = line[2:].strip()
                        if topic:
                            new_topics.append(topic)
        else:
            # Fallback: try to parse the whole content as YAML
            explanation = "Changes applied successfully."
            cleaned = _strip_code_fences(content)
            try:
                parsed = yaml.safe_load(cleaned)
                if isinstance(parsed, list):
                    new_topics = [str(t).strip() for t in parsed if t and str(t).strip()]
            except:
                pass
        
        # If we couldn't extract topics, return just the explanation
        if not new_topics:
            return jsonify({
                'success': True,
                'response': content,
                'new_topics': None
            })
        
        return jsonify({
            'success': True,
            'response': explanation if explanation else "ביצעתי את השינויים המבוקשים.",
            'new_topics': new_topics
        })
        
    except Exception as e:
        import traceback
        print(f"Chat feedback error: {e}\n{traceback.format_exc()}", file=sys.stderr)
        return jsonify({'error': str(e)}), 500


@app.route('/save_updated_topics', methods=['POST'])
def save_updated_topics():
    """Save updated topics and regenerate the output file."""
    try:
        data = request.json
        new_topics = data.get('topics', [])
        answer_column = data.get('answer_column', '')
        req_session_id = data.get('session_id')
        
        if not new_topics:
            return jsonify({'error': 'No topics provided'}), 400
        
        # Get file info from session or backup store
        file_info = get_file_info(req_session_id)
        if not file_info or not file_info.get('uploaded_file'):
            return jsonify({'error': 'No file uploaded. Please upload a file and generate topics again.'}), 400
        
        filepath = Path(file_info['uploaded_file'])
        current_sheet = file_info.get('current_sheet')
        original_filename = file_info.get('original_filename', 'output')
        
        if not filepath.exists():
            return jsonify({'error': 'Uploaded file not found. Please upload the file again.'}), 400
        
        # Load the data
        df = load_data_for_processing(filepath, current_sheet)
        answer_idx = column_letter_to_index(answer_column)
        
        if answer_idx >= len(df.columns):
            return jsonify({'error': 'Invalid answer column'}), 400
        
        answer_column_name = df.columns[answer_idx]
        
        # Remove any previously added topic columns FIRST (before checking for empty rows)
        old_topics = file_info.get('generated_topics', [])
        cols_to_keep = []
        for col in df.columns:
            if col not in old_topics:
                cols_to_keep.append(col)
        df = df[cols_to_keep]
        
        # NOW remove trailing empty rows (after removing old topic columns)
        def is_row_empty(row):
            for val in row:
                if pd.notna(val):
                    str_val = str(val).strip()
                    if str_val and str_val.lower() not in ('nan', 'none', ''):
                        return False
            return True
        
        while len(df) > 0:
            if is_row_empty(df.iloc[-1]):
                df = df.iloc[:-1]
            else:
                break
        
        # Insert new topic columns right after the answer column
        answer_col_position = df.columns.get_loc(answer_column_name) + 1
        for i, topic in enumerate(new_topics):
            if topic not in df.columns:
                df.insert(answer_col_position + i, topic, '')
        
        # Create the CSV content and save to temp file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        original_name = Path(original_filename).stem
        output_filename = f"{original_name}_with_topics_{timestamp}.csv"
        
        # Save to a temp file for quick download
        output_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='wb')
        output_temp.write(b'\xef\xbb\xbf')  # UTF-8 BOM for Excel
        csv_string = df.to_csv(index=False, encoding='utf-8')
        output_temp.write(csv_string.encode('utf-8'))
        output_temp.close()
        
        # Update both session and backup store
        sid = req_session_id or session.get('session_id')
        update_file_info(sid,
            generated_topics=new_topics,
            download_file_path=output_temp.name,
            download_filename=output_filename
        )
        
        return jsonify({
            'success': True,
            'message': 'Topics saved successfully',
            'file_ready': {
                'path': output_temp.name,
                'filename': output_filename
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Save topics error: {e}\n{traceback.format_exc()}", file=sys.stderr)
        return jsonify({'error': str(e)}), 500


def open_browser():
    url = 'http://127.0.0.1:5001'
    
    if platform.system() == 'Windows':
        chrome_paths = [
            r'C:\Program Files\Google\Chrome\Application\chrome.exe',
            r'C:\Program Files (x86)\Google\Chrome\Application\chrome.exe',
            os.path.expanduser(r'~\AppData\Local\Google\Chrome\Application\chrome.exe'),
        ]
        
        chrome_path = None
        for path in chrome_paths:
            if os.path.exists(path):
                chrome_path = path
                break
        
        if chrome_path:
            try:
                subprocess.Popen([chrome_path, url])
                return
            except Exception:
                pass
    
    webbrowser.open(url)


if __name__ == '__main__':
    Timer(1, open_browser).start()
    
    print("=" * 50)
    print("  Topic Generator Web App")
    print("=" * 50)
    print("Starting server...")
    print("Opening browser at http://127.0.0.1:5001")
    print("Press Ctrl+C to stop the server")
    print("=" * 50)
    print()
    
    try:
        app.run(debug=False, port=5001, host='127.0.0.1', threaded=True)
    except OSError as e:
        if "Address already in use" in str(e):
            print("\n❌ ERROR: Port 5001 is already in use!")
            sys.exit(1)
        else:
            raise

